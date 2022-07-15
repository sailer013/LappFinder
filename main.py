import openpyxl
import selenium.common.exceptions
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

class LappFinder:
    def __init__(self):
        self.service = Service("chromedriver.exe")
        self.options = webdriver.ChromeOptions()
        self.options.add_argument('headless')
        self.options.add_argument('window-size=1920x1080')
        self.options.add_argument("disable-gpu")

    def find_element_by_class(self, class_name):
        self.driver.implicitly_wait(2)
        element = self.driver.find_element(By.CLASS_NAME, class_name)
        return element

    def find_element_by_id(self, id):
        self.driver.implicitly_wait(2)
        element = self.driver.find_element(By.ID, id)
        return element

    def get_url_rs_online(self, part_number):
        product_url_list = []
        self.driver = webdriver.Chrome(service=self.service, options=self.options)
        self.driver.get('https://uk.rs-online.com/web/')
        accept_cookie_button = self.find_element_by_id('ensCloseBanner')
        accept_cookie_button.click()
        search_bar = self.find_element_by_id('searchBarTextInput')
        self.driver.implicitly_wait(3)
        search_bar.send_keys(part_number)
        self.driver.implicitly_wait(4)
        try:
            search_bar.send_keys(Keys.ENTER)
            category = self.driver.find_element(By.CSS_SELECTOR, "a[data-qa='category-button']")
            category.click()
            self.driver.implicitly_wait(3)
        except selenium.common.exceptions.WebDriverException:
            pass
        try:
            products = self.driver.find_elements(By.CSS_SELECTOR, "a[data-qa='product-tile-container']")
            for product in products:
                product_url_list.append(product.get_attribute('href'))
        except:
            products = []
        if not products:
            product_url_list.append(self.driver.current_url)
        self.driver.quit()
        return product_url_list, part_number

    def get_product_details(self, url_list, part_number):
        all_product_details_list = []
        if url_list:
            for link in url_list:
                print(part_number)
                product_details_list = []
                r = requests.get(link)
                if r'"Sorry, we couldn\'t find any results' in r.text:
                    all_product_details_list.append([part_number, 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'])
                    continue
                soup = BeautifulSoup(r.text, 'html.parser')
                product_details_list.append(part_number)
                try:
                    mfr_number = soup.select_one('dt:-soup-contains("Mfr") + dd').text
                    product_details_list.append(mfr_number)
                except:
                    product_details_list.append('N/A')
                name = soup.find('h1', attrs={'data-testid': 'long-description'}).text
                product_details_list.append(name)
                try:
                    stock_num = soup.select_one('dt:-soup-contains("RS") + dd').text
                    product_details_list.append(stock_num)
                except:
                    product_details_list.append('N/A')
                price = soup.find('p', attrs={'data-testid': 'price-exc-vat'}).text
                product_details_list.append(price)
                try:
                    volume = soup.find('p', attrs={'data-testid':'price-heading'}).text
                    product_details_list.append(volume)
                except:
                    product_details_list.append('N/A')
                try:
                    availability = soup.find('div', attrs={'data-testid': 'stock-status-0'}).text
                except:
                    availability = soup.find('div', attrs={'data-testid': 'stock-status-unknown'}).text
                product_details_list.append(availability)
                product_details_list.append(link)
                all_product_details_list.append(product_details_list)
        else:
            all_product_details_list.append([part_number, 'N/A', 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'])

        return all_product_details_list

    def load_values(self, file):
        with open(file) as file:
            values_list = []
            for row in file.readlines():
                values_list.append(row.strip())
        return values_list

    def clear_excel(self):
        wb = Workbook()
        ws = wb.active
        headers = ['Input', 'Nr', 'Name', 'Stock No.', 'Price', 'Volume', 'Avability', 'Link']
        ws.append(headers)
        wb.save('LappFinder.xlsx')

    def details_to_excel(self, data):
        wb = load_workbook('LappFinder.xlsx')
        ws = wb.active
        for row in data:
            for val in row:
                ws.append(val)
        wb.save('LappFinder.xlsx')

    def main(self):
        values = self.load_values('Input.txt')
        self.clear_excel()
        for value in values:
            data_list = []
            product_url, part_number = self.get_url_rs_online(value)
            details = self.get_product_details(product_url, part_number)
            data_list.append(details)
            self.details_to_excel(data_list)

if __name__ == '__main__':
    finder = LappFinder()
    finder.main()

